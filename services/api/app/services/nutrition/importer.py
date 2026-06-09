import csv
import gzip
import re
import unicodedata
from collections.abc import Iterator
from pathlib import Path
from typing import Iterable, TextIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import NutritionFoodReference


def first_value(row: dict[str, str], candidates: Iterable[str]) -> str | None:
    normalized_row = {normalize_column_name(key): value for key, value in row.items()}
    for candidate in candidates:
        value = row.get(candidate)
        if value not in (None, ""):
            return value
        value = normalized_row.get(normalize_column_name(candidate))
        if value not in (None, ""):
            return value
    return None


def normalize_column_name(value: str) -> str:
    without_accents = "".join(
        char for char in unicodedata.normalize("NFKD", value) if not unicodedata.combining(char)
    )
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", without_accents.lower())).strip()


def numeric_value(row: dict[str, str], candidates: Iterable[str]) -> float:
    return numeric_value_or_none(row, candidates) or 0.0


def numeric_value_or_none(row: dict[str, str], candidates: Iterable[str]) -> float | None:
    raw = first_value(row, candidates)
    if raw is None:
        return None
    normalized = raw.replace(",", ".").replace("\u00a0", "").strip()
    try:
        return float(normalized)
    except ValueError:
        return None


def open_food_facts_energy_kcal(row: dict[str, str]) -> float:
    kcal = numeric_value_or_none(row, ("energy-kcal_100g", "energy_kcal_100g"))
    if kcal is not None:
        return kcal
    kj = numeric_value_or_none(row, ("energy_100g", "energy-kj_100g", "energy_kj_100g"))
    if kj is None:
        return 0.0
    return kj / 4.184


def has_open_food_facts_nutrients(row: dict[str, str]) -> bool:
    return any(
        numeric_value_or_none(row, candidates) is not None
        for candidates in (
            ("energy-kcal_100g", "energy_kcal_100g", "energy_100g", "energy-kj_100g", "energy_kj_100g"),
            ("proteins_100g", "protein_g_100g"),
            ("carbohydrates_100g", "carbohydrates_g_100g"),
            ("fat_100g", "fat_g_100g"),
        )
    )


class NutritionDatasetImporter:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def import_ciqual_file(self, path: Path, dataset_version: str) -> int:
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            return await self.import_ciqual_xlsx(path, dataset_version)
        return await self.import_ciqual_csv(path, dataset_version)

    async def import_ciqual_csv(self, path: Path, dataset_version: str) -> int:
        return await self._import_ciqual_rows(self._rows(path), dataset_version)

    async def import_ciqual_xlsx(self, path: Path, dataset_version: str) -> int:
        return await self._import_ciqual_rows(self._xlsx_rows(path), dataset_version)

    async def _import_ciqual_rows(self, rows: Iterable[dict[str, str]], dataset_version: str) -> int:
        count = 0
        for row in rows:
            source_id = first_value(row, ("alim_code", "code", "source_id"))
            name = first_value(row, ("alim_nom_fr", "name", "alim_nom_eng"))
            if not source_id or not name:
                continue
            await self._upsert(
                source="ciqual",
                source_id=source_id,
                barcode=None,
                name=name,
                energy_kcal_100g=numeric_value(
                    row,
                    (
                        "energy_kcal_100g",
                        "Energie, Règlement UE N° 1169/2011 (kcal/100 g)",
                        "Energie (kcal/100 g)",
                    ),
                ),
                protein_g_100g=numeric_value(row, ("protein_g_100g", "Protéines, N x facteur de Jones (g/100 g)")),
                carbohydrates_g_100g=numeric_value(row, ("carbohydrates_g_100g", "Glucides (g/100 g)")),
                fat_g_100g=numeric_value(row, ("fat_g_100g", "Lipides (g/100 g)")),
                dataset_version=dataset_version,
                raw=row,
            )
            count += 1
        await self.db.flush()
        return count

    async def import_open_food_facts_csv(
        self,
        path: Path,
        dataset_version: str,
        *,
        country_tag: str | None = None,
        limit: int | None = None,
    ) -> int:
        count = 0
        for row in self._rows(path):
            if country_tag and not row_has_country_tag(row, country_tag):
                continue
            barcode = first_value(row, ("code", "barcode"))
            name = first_value(row, ("product_name_fr", "product_name", "generic_name_fr", "name"))
            if not barcode or not name:
                continue
            if not has_open_food_facts_nutrients(row):
                continue
            await self._upsert(
                source="openfoodfacts",
                source_id=barcode,
                barcode=barcode,
                name=name,
                energy_kcal_100g=open_food_facts_energy_kcal(row),
                protein_g_100g=numeric_value(row, ("proteins_100g", "protein_g_100g")),
                carbohydrates_g_100g=numeric_value(row, ("carbohydrates_100g", "carbohydrates_g_100g")),
                fat_g_100g=numeric_value(row, ("fat_100g", "fat_g_100g")),
                dataset_version=dataset_version,
                raw=row,
            )
            count += 1
            if limit is not None and count >= limit:
                break
        await self.db.flush()
        return count

    @staticmethod
    def _rows(path: Path) -> Iterator[dict[str, str]]:
        opener = gzip.open if path.suffix == ".gz" else Path.open
        with opener(path, "rt", encoding="utf-8-sig", newline="") as handle:
            yield from csv.DictReader(handle, dialect=detect_csv_dialect(handle))

    @staticmethod
    def _xlsx_rows(path: Path) -> Iterator[dict[str, str]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["composition nutritionnelle"] if "composition nutritionnelle" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows, ())]
        for values in rows:
            yield {
                header: "" if value is None else str(value).strip()
                for header, value in zip(headers, values, strict=False)
                if header
            }

    async def _upsert(
        self,
        *,
        source: str,
        source_id: str,
        barcode: str | None,
        name: str,
        energy_kcal_100g: float,
        protein_g_100g: float,
        carbohydrates_g_100g: float,
        fat_g_100g: float,
        dataset_version: str,
        raw: dict,
    ) -> NutritionFoodReference:
        reference = await self.db.scalar(
            select(NutritionFoodReference).where(
                NutritionFoodReference.source == source,
                NutritionFoodReference.source_id == source_id,
            )
        )
        if reference is None:
            reference = NutritionFoodReference(source=source, source_id=source_id, name=name, dataset_version=dataset_version)
            self.db.add(reference)
        reference.barcode = barcode
        reference.name = name
        reference.energy_kcal_100g = energy_kcal_100g
        reference.protein_g_100g = protein_g_100g
        reference.carbohydrates_g_100g = carbohydrates_g_100g
        reference.fat_g_100g = fat_g_100g
        reference.dataset_version = dataset_version
        reference.raw_json = raw
        return reference


def detect_csv_dialect(handle: TextIO) -> csv.Dialect:
    sample = handle.read(8192)
    handle.seek(0)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        return csv.get_dialect("excel")


def row_has_country_tag(row: dict[str, str], country_tag: str) -> bool:
    tags = first_value(row, ("countries_tags", "countries")) or ""
    normalized_country = country_tag.strip().lower()
    return normalized_country in {tag.strip().lower() for tag in tags.split(",")}
